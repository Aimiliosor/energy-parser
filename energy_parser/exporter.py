import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from rich.console import Console

console = Console()


def save_xlsx(df: pd.DataFrame, output_path: str, is_final: bool = False):
    """Save DataFrame to formatted XLSX file.

    Args:
        df: DataFrame with 'Date & Time' and value columns
        output_path: Full path for the output file
        is_final: If True, apply full formatting (final output)
    """
    label = "Final" if is_final else "Initial"
    console.print(f"\n  Saving {label.lower()} XLSX: [bold]{os.path.basename(output_path)}[/bold]")

    # Format Date & Time as string for consistent display
    df_out = df.copy()
    df_out["Date & Time"] = df_out["Date & Time"].dt.strftime("%Y-%m-%d %H:%M")

    # Write with pandas first
    df_out.to_excel(output_path, index=False, engine="openpyxl")

    # Now open with openpyxl for formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Header formatting
    header_font = Font(bold=True, size=11)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header = ws.cell(row=1, column=col_idx).value or ""

        if "Date" in str(header):
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = 18

    # Number formatting for value columns (2 decimal places)
    if is_final:
        for col_idx in range(2, ws.max_column + 1):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

    # Date column alignment
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal="left")

    wb.save(output_path)
    console.print(f"  [green]Saved: {output_path}[/green]")
    console.print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")


def save_cost_simulation_xlsx(output_path: str,
                               summary_data: dict,
                               monthly_data: dict,
                               comparison_data: list[dict] | None = None):
    """Save cost simulation results to a formatted XLSX file.

    Args:
        output_path: Full path for the output file
        summary_data: Dict with CostBreakdown fields for the overall summary
        monthly_data: Dict of {month_key: dict with CostBreakdown fields}
        comparison_data: Optional list of scenario comparison dicts
    """
    console.print(f"\n  Saving cost simulation XLSX: "
                  f"[bold]{os.path.basename(output_path)}[/bold]")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_rows = [
            {"Metric": "Total Consumption (MWh)",
             "Value": round(summary_data.get("total_consumption_kwh", 0) / 1000, 1)},
            {"Metric": "Total Production (MWh)",
             "Value": round(summary_data.get("total_production_kwh", 0) / 1000, 1)},
            {"Metric": "Self-Consumed (MWh)",
             "Value": round(summary_data.get("self_consumed_kwh", 0) / 1000, 1)},
            {"Metric": "Grid Consumed (MWh)",
             "Value": round(summary_data.get("grid_consumed_kwh", 0) / 1000, 1)},
            {"Metric": "Injected (MWh)",
             "Value": round(summary_data.get("injected_kwh", 0) / 1000, 1)},
            {"Metric": "Self-Consumption Rate",
             "Value": f"{summary_data.get('self_consumption_rate', 0):.1%}"},
            {"Metric": "Autarky Rate",
             "Value": f"{summary_data.get('autarky_rate', 0):.1%}"},
            {"Metric": "Energy Cost (\u20ac)",
             "Value": round(summary_data.get("energy_cost", 0), 2)},
            {"Metric": "Grid Capacity Cost (\u20ac)",
             "Value": round(summary_data.get("grid_capacity_cost", 0), 2)},
            {"Metric": "Grid Energy Cost (\u20ac)",
             "Value": round(summary_data.get("grid_energy_cost", 0), 2)},
            {"Metric": "Taxes & Levies (\u20ac)",
             "Value": round(summary_data.get("taxes_and_levies", 0), 2)},
            {"Metric": "Overshoot Penalties (\u20ac)",
             "Value": round(summary_data.get("overshoot_penalties", 0), 2)},
            {"Metric": "Injection Revenue (\u20ac)",
             "Value": round(summary_data.get("injection_revenue", 0), 2)},
            {"Metric": "Total Cost excl. VAT (\u20ac)",
             "Value": round(summary_data.get("total_cost_excl_vat", 0), 2)},
            {"Metric": "Peak Demand (kW)",
             "Value": round(summary_data.get("peak_demand_kw", 0), 1)},
            {"Metric": "Overshoots",
             "Value": summary_data.get("overshoots_count", 0)},
        ]
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Cost Summary",
                            index=False)

        # Monthly sheet
        if monthly_data:
            monthly_rows = []
            for month_key in sorted(monthly_data.keys()):
                md = monthly_data[month_key]
                cons_kwh = md.get("total_consumption_kwh", 0)
                total = md.get("total_cost_excl_vat", 0)
                monthly_rows.append({
                    "Month": month_key,
                    "Consumption (MWh)": round(cons_kwh / 1000, 1),
                    "Production (MWh)": round(
                        md.get("total_production_kwh", 0) / 1000, 1),
                    "Self-Cons Rate": f"{md.get('self_consumption_rate', 0):.0%}",
                    "Energy Cost (\u20ac)": round(
                        md.get("energy_cost", 0), 2),
                    "Grid Cost (\u20ac)": round(
                        md.get("grid_capacity_cost", 0)
                        + md.get("grid_energy_cost", 0), 2),
                    "Taxes (\u20ac)": round(
                        md.get("taxes_and_levies", 0), 2),
                    "Penalties (\u20ac)": round(
                        md.get("overshoot_penalties", 0), 2),
                    "Total Cost (\u20ac)": round(total, 2),
                    "Avg (\u20ac/kWh)": round(
                        total / max(cons_kwh, 1), 4),
                    "Peak Demand (kW)": round(
                        md.get("peak_demand_kw", 0), 1),
                })
            df_monthly = pd.DataFrame(monthly_rows)
            df_monthly.to_excel(writer, sheet_name="Monthly Breakdown",
                                index=False)

        # Comparison sheet
        if comparison_data:
            df_comp = pd.DataFrame(comparison_data)
            df_comp.to_excel(writer, sheet_name="Scenario Comparison",
                             index=False)

    # Format with openpyxl
    wb = load_workbook(output_path)
    header_font = Font(bold=True, size=11)
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

    wb.save(output_path)
    console.print(f"  [green]Saved: {output_path}[/green]")
